import openpyxl
import os

src = r"C:\Users\a8466\Desktop\工作文件\客户导入模板_最终版.xlsx"
wb = openpyxl.load_workbook(src)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print("Headers:", headers)
print("Total data rows:", ws.max_row - 1)

# Find country column
country_col = None
for i, h in enumerate(headers):
    if h:
        hs = str(h).strip()
        if "国家" in hs or "地区" in hs or "country" in hs.lower() or "region" in hs.lower():
            country_col = i
            break

print("Country column index:", country_col, "->", headers[country_col] if country_col is not None else "NOT FOUND")

# Collect all countries and Indian rows
all_countries = set()
india_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    val = row[country_col] if country_col is not None else None
    if val:
        all_countries.add(str(val).strip())
    if val and ("印度" in str(val).strip() or "india" in str(val).strip().lower()):
        india_rows.append(row)

print(f"\nIndia customers: {len(india_rows)}")
print("All countries found:", sorted(all_countries))

# Preview first India row
if india_rows:
    print("\nFirst India row preview:")
    for h, v in zip(headers, india_rows[0]):
        print(f"  {h}: {v}")

# Export to new Excel
if india_rows:
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "印度客户"
    out_ws.append(headers)
    for r in india_rows:
        out_ws.append(list(r))
    out = r"C:\Users\a8466\Desktop\工作文件\印度客户.xlsx"
    out_wb.save(out)
    print(f"\nExported to: {out}")
else:
    print("\nNo Indian customers found.")
