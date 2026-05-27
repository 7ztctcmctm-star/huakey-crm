#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel名片导入CRM系统脚本
读取Excel文件，将客户数据导入到huakey_crm数据库的crm_customer表
"""

import os
import sys
import logging
from datetime import datetime

import pymysql
import openpyxl

# 脚本所在目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 配置常规日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.StreamHandler(sys.stdout),
    ]
)

# 错误日志 - 记录到 error.log
error_log_path = os.path.join(SCRIPT_DIR, 'error.log')
error_logger = logging.getLogger('error')
error_logger.setLevel(logging.ERROR)
if not error_logger.handlers:
    try:
        error_handler = logging.FileHandler(error_log_path, encoding='utf-8', mode='a')
        error_handler.setFormatter(logging.Formatter('%(asctime)s [ERROR] %(message)s', '%Y-%m-%d %H:%M:%S'))
        error_logger.addHandler(error_handler)
    except Exception:
        # 如果无法创建日志文件，使用控制台输出
        console = logging.StreamHandler(sys.stderr)
        console.setFormatter(logging.Formatter('%(asctime)s [ERROR] %(message)s', '%Y-%m-%d %H:%M:%S'))
        error_logger.addHandler(console)
    error_logger.propagate = False

# ============ 配置区域 ============

DB_CONFIG = {
    'host': 'localhost',
    'port': 3306,
    'user': 'crm_user',
    'password': 'Huakey@2024',
    'database': 'huakey_crm',
    'charset': 'utf8mb4'
}

# Excel列名到数据库字段的映射
# key: Excel中可能出现的列名（支持多种写法）
# value: 对应的数据库字段名
COLUMN_MAPPING = {
    # 公司名称
    '公司名称': 'company_name',
    '公司名': 'company_name',
    '企业名称': 'company_name',
    '单位名称': 'company_name',
    '客户名称': 'company_name',
    # 联系人
    '联系人': 'contact_name',
    '联系人姓名': 'contact_name',
    '姓名': 'contact_name',
    '联系人名': 'contact_name',
    # 电话
    '电话': 'phone',
    '联系电话': 'phone',
    '手机': 'phone',
    '电话号码': 'phone',
    '联系电话': 'phone',
    # 邮箱
    '邮箱': 'email',
    '电子邮件': 'email',
    '邮件': 'email',
    'E-mail': 'email',
    'Email': 'email',
    # 地址
    '地址': 'address',
    '公司地址': 'address',
    '详细地址': 'address',
    # 行业
    '行业': 'industry',
    '所属行业': 'industry',
    '行业类型': 'industry',
    # 来源
    '来源': 'source',
    '客户来源': 'source',
    '渠道': 'source',
    # 等级
    '等级': 'level',
    '客户等级': 'level',
    # 备注
    '备注': 'remark',
    '说明': 'remark',
    '描述': 'remark',
}


# ============ 核心函数 ============

def read_excel(file_path):
    """读取Excel文件，返回 (表头列表, 数据行列表)"""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.xlsx':
        return read_xlsx(file_path)
    elif ext == '.xls':
        return read_xls(file_path)
    else:
        raise ValueError(f'不支持的文件格式: {ext}，请使用 .xlsx 或 .xls 文件')


def read_xlsx(file_path):
    """读取 .xlsx 文件 (openpyxl)"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError('Excel文件为空')

    headers = [str(h).strip() if h else '' for h in rows[0]]
    data = []
    for row in rows[1:]:
        data.append([str(c).strip() if c is not None else '' for c in row])

    return headers, data


def read_xls(file_path):
    """读取 .xls 文件 (xlrd)"""
    import xlrd

    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)

    if ws.nrows == 0:
        raise ValueError('Excel文件为空')

    headers = [str(ws.cell_value(0, col)).strip() for col in range(ws.ncols)]
    data = []
    for row_idx in range(1, ws.nrows):
        row_data = []
        for col_idx in range(ws.ncols):
            cell = ws.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                # 处理日期类型
                dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                row_data.append(dt.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                row_data.append(str(cell.value).strip() if cell.value != '' else '')
        data.append(row_data)

    return headers, data


def map_columns(headers):
    """
    将Excel表头映射为数据库字段名
    返回: {
        column_index: ('column_name', 'db_field_name'),
        ...
    }
    """
    mapping = {}
    for idx, header in enumerate(headers):
        if header in COLUMN_MAPPING:
            mapping[idx] = (header, COLUMN_MAPPING[header])

    if not mapping:
        logging.warning('未能识别任何列名，支持的列名: %s', list(COLUMN_MAPPING.keys()))

    return mapping


def build_records(data_rows, column_mapping):
    """将Excel数据行转换为数据库记录列表"""
    records = []
    for row_idx, row in enumerate(data_rows):
        record = {}
        for col_idx, (excel_col, db_col) in column_mapping.items():
            if col_idx < len(row) and row[col_idx]:
                record[db_col] = row[col_idx]

        # 跳过完全空白的行
        if not any(v for v in record.values()):
            continue

        records.append(record)

    return records


def connect_db():
    """连接MySQL数据库"""
    try:
        conn = pymysql.connect(**DB_CONFIG)
        logging.info('数据库连接成功')
        return conn
    except Exception as e:
        logging.error('数据库连接失败: %s', e)
        sys.exit(1)


def check_duplicate(cursor, record):
    """
    检查是否重复
    根据公司名称或电话判断是否已存在
    返回: (is_duplicate, existing_company_name)
    """
    company_name = record.get('company_name', '').strip()
    phone = record.get('phone', '').strip()

    if not company_name and not phone:
        return False, None

    conditions = []
    params = []

    if company_name:
        conditions.append('company_name = %s')
        params.append(company_name)

    if phone:
        conditions.append('phone = %s')
        params.append(phone)

    where = ' OR '.join(conditions)
    sql = f'SELECT id, company_name, phone FROM crm_customer WHERE status != 0 AND ({where}) LIMIT 1'
    cursor.execute(sql, params)
    result = cursor.fetchone()

    if result:
        return True, result[1]

    return False, None


def show_preview(records, column_mapping):
    """显示前5条数据预览"""
    print('\n' + '=' * 80)
    print('  📋 数据预览（前5条）')
    print('=' * 80)

    # 获取所有出现的字段和对应的Excel列名
    excel_names = {}
    for col_idx, (excel_col, db_col) in column_mapping.items():
        excel_names[db_col] = excel_col

    preview_count = min(5, len(records))
    for i in range(preview_count):
        record = records[i]
        print(f'\n  第 {i + 1} 条:')
        for db_col in ['company_name', 'contact_name', 'phone', 'email',
                        'address', 'industry', 'source', 'level']:
            value = record.get(db_col, '')
            if value:
                excel_name = excel_names.get(db_col, db_col)
                print(f'    {excel_name}: {value}')
        if record.get('remark'):
            print(f'    备注: {record["remark"]}')

    if len(records) > 5:
        print(f'\n  ... 还有 {len(records) - 5} 条数据未显示')

    print('=' * 80)


def import_data(conn, records, skip_duplicates=True):
    """
    将数据导入数据库
    skip_duplicates: 是否跳过重复数据
    返回: (success_count, fail_count, skip_count)
    """
    cursor = conn.cursor()
    success_count = 0
    fail_count = 0
    skip_count = 0

    total = len(records)

    for idx, record in enumerate(records):
        row_num = idx + 1
        try:
            # 检查必填字段
            if not record.get('company_name'):
                error_logger.error(f'第{row_num}条: 公司名称为空，跳过')
                skip_count += 1
                continue

            # 检查重复
            if skip_duplicates:
                is_dup, existing_name = check_duplicate(cursor, record)
                if is_dup:
                    error_logger.error(
                        f'第{row_num}条: 重复数据，已存在公司"{existing_name}"'
                        f'（Excel: "{record.get("company_name")}"）'
                    )
                    skip_count += 1
                    continue

            # 构建插入SQL
            fields = ['company_name', 'contact_name', 'phone', 'email',
                      'address', 'industry', 'source', 'level', 'remark']

            # 检查 source 是否合法（细化后的来源分类）
            source = record.get('source')
            valid_sources = {
                '展会',
                'Facebook', 'Instagram', 'LinkedIn', '独立站', '其他网络渠道',
                '转介绍',
                '电话',
                '其他'
            }
            # 兼容旧数据：'网络' → '其他网络渠道'
            source_compat_map = {
                '网络': '其他网络渠道'
            }
            if source and source not in valid_sources:
                # 兼容映射
                if source in source_compat_map:
                    record['source'] = source_compat_map[source]
                    logging.info(f'第{row_num}条: 来源"{source}"已自动映射为"{source_compat_map[source]}"')
                else:
                    # 尝试模糊匹配
                    matched = False
                    for vs in valid_sources:
                        if source.lower() == vs.lower():
                            record['source'] = vs
                            matched = True
                            break
                    if not matched:
                        record['source'] = '其他'
                        logging.warning(f'第{row_num}条: 来源"{source}"不在可选范围，设为"其他"')

            # 检查 level 是否合法
            level = record.get('level', '')
            valid_levels = {'A', 'B', 'C', 'D'}
            if level and level.upper() in valid_levels:
                record['level'] = level.upper()
            elif level and level not in valid_levels:
                record['level'] = 'C'
                logging.warning(f'第{row_num}条: 等级"{level}"不在可选范围，设为"C"')

            # 重新构建参数（因为可能修改了 source/level）
            fields_list = []
            values_list = []
            for field in fields:
                val = record.get(field, None)
                fields_list.append(field)
                values_list.append(val if val else None)

            placeholders = ', '.join(['%s'] * len(fields_list))
            sql = f'INSERT INTO crm_customer ({", ".join(fields_list)}) VALUES ({placeholders})'

            cursor.execute(sql, values_list)
            conn.commit()
            success_count += 1

        except pymysql.Error as e:
            conn.rollback()
            error_msg = f'第{row_num}条: 数据库错误 - {e}'
            error_logger.error(error_msg)
            fail_count += 1

        except Exception as e:
            conn.rollback()
            error_msg = f'第{row_num}条: {e}'
            error_logger.error(error_msg)
            fail_count += 1

        # 每100条打印进度
        if (idx + 1) % 100 == 0:
            logging.info(f'  进度: {idx + 1}/{total}')

    cursor.close()
    return success_count, fail_count, skip_count


def main():
    """主函数"""
    print('=' * 60)
    print('  📊 Excel名片导入CRM系统')
    print('  铧旗CRM - 客户数据导入工具')
    print('=' * 60)

    # 解析命令行参数
    auto_confirm = '--yes' in sys.argv or '-y' in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith('-')]

    # 1. 获取文件路径
    if len(args) > 0:
        file_path = args[0]
    else:
        file_path = input('\n请输入Excel文件路径: ').strip().strip('"').strip("'")

    if not os.path.exists(file_path):
        print(f'\n❌ 文件不存在: {file_path}')
        sys.exit(1)

    print(f'\n📄 文件: {os.path.basename(file_path)}')
    print(f'📁 路径: {os.path.dirname(file_path) or "当前目录"}')

    # 2. 读取Excel
    try:
        headers, data_rows = read_excel(file_path)
    except Exception as e:
        print(f'\n❌ 读取Excel失败: {e}')
        sys.exit(1)

    print(f'\n✅ 读取成功: {len(data_rows)} 行数据, {len(headers)} 列')
    print(f'   表头: {", ".join(headers)}')

    # 3. 映射列名
    column_mapping = map_columns(headers)
    if not column_mapping:
        print('\n❌ 未能识别任何有效列名，请检查Excel表头是否正确')
        print('   支持的列名:', ', '.join(COLUMN_MAPPING.keys()))
        sys.exit(1)

    mapped_fields = set(v[1] for v in column_mapping.values())
    print(f'   识别到 {len(column_mapping)} 个字段映射: {", ".join(mapped_fields)}')

    # 4. 转换数据
    records = build_records(data_rows, column_mapping)

    if not records:
        print('\n❌ 没有有效数据可导入（所有行均为空）')
        sys.exit(1)

    print(f'   有效数据: {len(records)} 条')

    # 5. 显示预览
    show_preview(records, column_mapping)

    # 6. 连接数据库
    print('\n🔗 正在连接数据库...')
    conn = connect_db()

    # 7. 检查重复
    print('\n🔍 正在检查重复数据...')
    cursor = conn.cursor()
    dup_count = 0
    clean_records = []
    for record in records:
        is_dup, existing = check_duplicate(cursor, record)
        if is_dup:
            dup_count += 1
        else:
            clean_records.append(record)
    cursor.close()

    if dup_count > 0:
        print(f'   发现 {dup_count} 条重复数据（将跳过）')
    print(f'   待导入: {len(clean_records)} 条')

    # 8. 确认导入
    print('\n' + '-' * 60)
    if len(clean_records) == 0:
        print('⚠️  没有新数据需要导入')
        conn.close()
        sys.exit(0)

    confirm = 'y' if auto_confirm else input(f'\n确认导入 {len(clean_records)} 条数据？(y/n): ').strip().lower()
    if confirm not in ('y', 'yes', '是'):
        print('已取消导入')
        conn.close()
        sys.exit(0)

    # 9. 执行导入
    print(f'\n🚀 开始导入 {len(clean_records)} 条数据...')
    start_time = datetime.now()

    success, fail, skip = import_data(conn, clean_records, skip_duplicates=False)

    elapsed = (datetime.now() - start_time).total_seconds()

    # 10. 显示统计
    print('\n' + '=' * 60)
    print('  📊 导入结果统计')
    print('=' * 60)
    print(f'  ✅ 成功导入: {success} 条')
    print(f'  ⏭️  重复跳过: {dup_count} 条')
    print(f'  ❌ 导入失败: {fail} 条')
    print(f'  ⏱️  耗时: {elapsed:.2f} 秒')

    if fail > 0:
        print(f'\n  ⚠️  失败记录已保存到 error.log')

    print('=' * 60)

    conn.close()


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print('\n\n⚠️  用户中断操作')
        sys.exit(0)
    except Exception as e:
        print(f'\n❌ 程序异常: {e}')
        logging.exception('程序异常')
        sys.exit(1)
